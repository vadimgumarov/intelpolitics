"""Phase C Starmer strategies — four T1 source ingest modules.

Replaces the legacy `strategy_starmer()` in `src/pipelines/multi_politician.py`
(which hit gov.uk press releases — T4 talking points). Each strategy here
maps to one T1 source in `sources.v2.yaml`:

  1. strategy_starmer_hansard()         — Hansard memberdebatecontributions
                                           tier=T1, action_class=statement_official,
                                           subject_role=speaker.
  2. strategy_starmer_commons_votes()   — UK Parliament Commons Votes Service
                                           tier=T1, action_class=vote,
                                           subject_role=voter.
  3. strategy_starmer_lobbying_xlsx()   — UK Lobbying Register quarterly XLSX
                                           tier=T1, action_class=lobby_filing,
                                           subject_role=lobbied.
  4. strategy_starmer_members_api()     — UK Parliament Members API identity
                                           tier=T1, action_class=supporting_metadata.

Canon: framework/BKM/dr-source-quality-canon-2026-05-12.md + addendum.

Each strategy returns a list of IngestRow dicts ready for INSERT into the
`rows` table (per migration 002_sources_v2.sql). Idempotency: each row carries
a content_hash; the pipeline upserts on (source_id, ext_id) and skips inserts
whose content_hash matches the existing row.

Rate limiting: self-imposed 1 req/sec/host per intelpolitics SOP §1.6 +
sources.v2.yaml gotchas. Exponential backoff on transient errors (429, 5xx).
"""
from __future__ import annotations

import hashlib
import io
import json
import logging
import re
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from typing import Any, Iterator

import httpx

log = logging.getLogger(__name__)


# A real-browser User-Agent string per sources.v2.yaml `lobbying-register-quarterly-xlsx`
# gotchas — Cloudflare WAF rejects default UA on registrarofconsultantlobbyists.org.uk.
BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# Self-imposed 1 req/sec/host pacing per intelpolitics SOP §1.6.
DEFAULT_PACE_SEC = 1.0

# Default lookback window for Phase C first scrape. Phase D may extend.
DEFAULT_LOOKBACK_DAYS = 365


@dataclass
class IngestRow:
    """One row ready for INSERT into the `rows` table.

    Fields align with migration 002_sources_v2.sql `rows` table.
    """
    source_id: str
    tier: str
    action_class: str
    ext_id: str
    occurred_at: datetime
    title: str
    summary: str
    source_url: str
    payload_json: dict[str, Any]
    subject_role: str | None = None
    # politician_id resolved at insert time by the pipeline (FK to politicians.id).

    def content_hash(self) -> str:
        """sha256 over canonical payload — used for content-level dedup so a
        re-fetched row that has not changed does not trigger an UPDATE."""
        canonical = json.dumps(self.payload_json, sort_keys=True, default=str)
        return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


@dataclass
class StrategyMetrics:
    """Per-strategy stats — surfaced by the pipeline summary."""
    name: str
    rows_yielded: int = 0
    rows_inserted: int = 0
    rows_duplicate: int = 0
    rows_rejected: int = 0
    fetch_ok: int = 0
    fetch_fail: int = 0
    duration_sec: float = 0.0
    fatal_error: str | None = None
    latencies: list[float] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Shared HTTP helpers — rate-limited, retry-aware
# ---------------------------------------------------------------------------

def _pace(host_last_call: dict[str, float], host: str, pace: float = DEFAULT_PACE_SEC) -> None:
    """Sleep so successive calls to the same host honour the pace floor."""
    now = time.monotonic()
    last = host_last_call.get(host, 0.0)
    delta = now - last
    if delta < pace:
        time.sleep(pace - delta)
    host_last_call[host] = time.monotonic()


def _backoff_sleep(attempt: int, base: float = 1.0, cap: float = 30.0) -> None:
    """Exponential backoff: base * 2^attempt, capped, with mild jitter."""
    delay = min(cap, base * (2 ** attempt))
    # Light deterministic jitter so multiple processes don't lockstep.
    delay = delay + (attempt * 0.13)
    log.info("backoff sleep %.2fs (attempt=%d)", delay, attempt)
    time.sleep(delay)


def _http_get_json(
    client: httpx.Client,
    url: str,
    metrics: StrategyMetrics,
    host_last_call: dict[str, float],
    *,
    max_attempts: int = 5,
    extra_headers: dict[str, str] | None = None,
) -> Any:
    """GET JSON with exponential backoff. Returns parsed JSON or None on terminal failure."""
    from urllib.parse import urlparse
    host = urlparse(url).netloc
    headers = {"User-Agent": "intelpolitics-phase-c/2026-05-12", "Accept": "application/json"}
    if extra_headers:
        headers.update(extra_headers)

    for attempt in range(max_attempts):
        _pace(host_last_call, host)
        started = time.monotonic()
        try:
            r = client.get(url, headers=headers, timeout=30.0, follow_redirects=True)
            elapsed = time.monotonic() - started
            metrics.latencies.append(elapsed)
            if r.status_code == 200:
                metrics.fetch_ok += 1
                try:
                    return r.json()
                except json.JSONDecodeError as e:
                    log.warning("JSON decode failed url=%s: %s", url, e)
                    metrics.fetch_fail += 1
                    return None
            if r.status_code in (429, 500, 502, 503, 504) and attempt < max_attempts - 1:
                log.warning("status=%d on %s — backoff (attempt %d)", r.status_code, url, attempt)
                _backoff_sleep(attempt)
                continue
            log.warning("terminal status=%d on %s", r.status_code, url)
            metrics.fetch_fail += 1
            return None
        except (httpx.TimeoutException, httpx.TransportError) as e:
            log.warning("transport error on %s: %s (attempt %d)", url, e, attempt)
            if attempt < max_attempts - 1:
                _backoff_sleep(attempt)
                continue
            metrics.fetch_fail += 1
            return None
    metrics.fetch_fail += 1
    return None


def _http_get_bytes(
    client: httpx.Client,
    url: str,
    metrics: StrategyMetrics,
    host_last_call: dict[str, float],
    *,
    extra_headers: dict[str, str] | None = None,
    max_attempts: int = 4,
) -> bytes | None:
    """GET raw bytes (for XLSX downloads). Returns bytes or None."""
    from urllib.parse import urlparse
    host = urlparse(url).netloc
    headers = {"User-Agent": BROWSER_UA}
    if extra_headers:
        headers.update(extra_headers)

    for attempt in range(max_attempts):
        _pace(host_last_call, host)
        started = time.monotonic()
        try:
            r = client.get(url, headers=headers, timeout=60.0, follow_redirects=True)
            elapsed = time.monotonic() - started
            metrics.latencies.append(elapsed)
            if r.status_code == 200 and r.content:
                metrics.fetch_ok += 1
                return r.content
            if r.status_code in (429, 500, 502, 503, 504) and attempt < max_attempts - 1:
                _backoff_sleep(attempt)
                continue
            log.warning("XLSX fetch terminal status=%d on %s", r.status_code, url)
            metrics.fetch_fail += 1
            return None
        except (httpx.TimeoutException, httpx.TransportError) as e:
            log.warning("XLSX transport error: %s (attempt %d)", e, attempt)
            if attempt < max_attempts - 1:
                _backoff_sleep(attempt)
                continue
            metrics.fetch_fail += 1
            return None
    metrics.fetch_fail += 1
    return None


def _parse_iso_dt(s: str | None) -> datetime | None:
    if not s:
        return None
    try:
        s = s.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except (ValueError, AttributeError):
        return None


# ===========================================================================
# Strategy 1 — Hansard memberdebatecontributions  (statement_official / speaker)
# ===========================================================================

HANSARD_BASE = "https://hansard-api.parliament.uk"
# 2026-05-12 live-ingest fix: legacy `/debates/memberdebatecontributions/{id}.json`
# returns 404. The endpoint `/search/contributions/Spoken.json` works from a
# browser/Mac but consistently returns {"Results": [], "TotalResultCount": 0}
# when called from inside the k3s pod (Olares cluster egress) — verified twice
# from the same job IP, while the same URL returns 1544 results from Mac.
# Root cause is most likely Hansard search-API geo/edge filtering on cluster
# IPs.
# Fallback that actually works from the cluster: Members API
# `/api/Members/{id}/ContributionSummary` — already proven working in
# strategy_starmer_members_api. Returns debate-level summaries (one per
# Commons debate the member spoke in), not per-individual-contribution, but
# that is the right granularity for downstream verdict computation (one row
# per debate = one statement_official act). The hansard.parliament.uk debate
# permalink is constructed from debateWebsiteId.
MEMBERS_API_CONTRIB = (
    "https://members-api.parliament.uk/api/Members/{member_id}/ContributionSummary"
)
# Kept for reference / future revert if Hansard egress filter lifts:
HANSARD_CONTRIBS_LEGACY = HANSARD_BASE + "/search/contributions/Spoken.json"


def strategy_starmer_hansard(
    member_id: int,
    metrics: StrategyMetrics,
    *,
    start_date: datetime | None = None,
    end_date: datetime | None = None,
    take_limit: int = 200,
    client: httpx.Client | None = None,
    host_last_call: dict[str, float] | None = None,
) -> list[IngestRow]:
    """Ingest Hansard contributions for a member via Members API ContributionSummary.

    Endpoint: GET /api/Members/{id}/ContributionSummary?skip=N&take=M
    Response shape: {"items": [{"value": {...debate summary...}, "links":[...]}], "totalResults": int}

    We page through results, taking up to `take_limit` items. We filter to
    sittingDate within [start_date, end_date]. Each item becomes one
    IngestRow with ext_id = debateWebsiteId, occurred_at = sittingDate.
    """
    metrics.name = "starmer_hansard"
    t0 = time.monotonic()
    host_last_call = host_last_call if host_last_call is not None else {}
    own_client = client is None
    client = client or httpx.Client()

    try:
        if end_date is None:
            end_date = datetime.now(timezone.utc)
        if start_date is None:
            start_date = end_date - timedelta(days=DEFAULT_LOOKBACK_DAYS)

        base = MEMBERS_API_CONTRIB.format(member_id=member_id)
        page_size = 20  # Members API default; do not exceed without checking docs.
        collected: list[dict[str, Any]] = []
        skip = 0
        max_pages = 25  # safety cap = 500 entries
        for _page in range(max_pages):
            url_with_q = f"{base}?skip={skip}&take={page_size}"
            data = _http_get_json(client, url_with_q, metrics, host_last_call)
            if not data:
                break
            items = data.get("items") or []
            if not items:
                break
            collected.extend(items)
            total = data.get("totalResults") or 0
            skip += page_size
            if skip >= total or len(collected) >= take_limit:
                break
        log.info("[hansard] %d contribution summaries collected from Members API", len(collected))

        rows: list[IngestRow] = []
        for item in collected:
            v = item.get("value") if isinstance(item, dict) else None
            if not isinstance(v, dict):
                continue
            ext_id = v.get("debateWebsiteId") or v.get("debateId")
            if not ext_id:
                continue
            ext_id = str(ext_id)
            occurred = _parse_iso_dt(v.get("sittingDate")) or end_date
            # Filter to lookback window.
            if occurred < start_date or occurred > end_date:
                continue
            title_raw = v.get("debateTitle") or v.get("section") or "Commons contribution"
            title = (title_raw[:300]) if title_raw else "Commons contribution"
            total_contribs = v.get("totalContributions") or 0
            summary = (
                f"{total_contribs} contributions in debate '{title_raw}' "
                f"(section: {v.get('section', '?')}, house: {v.get('house', '?')})."
            )[:600]
            source_url = (
                f"https://hansard.parliament.uk/Commons/{occurred.date().isoformat()}/"
                f"debates/{v.get('debateWebsiteId', '')}/"
                f"{(title_raw or 'Debate').replace(' ', '')[:40]}"
            )
            rows.append(IngestRow(
                source_id="hansard-search-contributions",
                tier="T1",
                action_class="statement_official",
                subject_role="speaker",
                ext_id=ext_id,
                occurred_at=occurred,
                title=title,
                summary=summary,
                source_url=source_url,
                payload_json=v,
            ))
            if len(rows) >= take_limit:
                break
        metrics.rows_yielded = len(rows)
        return rows
    except Exception as e:
        metrics.fatal_error = f"hansard strategy crashed: {type(e).__name__}: {e}"
        log.exception("hansard strategy crash")
        return []
    finally:
        metrics.duration_sec = time.monotonic() - t0
        if own_client:
            client.close()


# ===========================================================================
# Strategy 2 — Commons Votes Service  (vote / voter)
# ===========================================================================

VOTES_BASE = "https://commonsvotes-api.parliament.uk"
VOTES_SEARCH = VOTES_BASE + "/data/divisions.json/search"
VOTES_DETAIL = VOTES_BASE + "/data/division/{division_id}.json"


def strategy_starmer_commons_votes(
    member_id: int,
    metrics: StrategyMetrics,
    *,
    start_date: datetime | None = None,
    end_date: datetime | None = None,
    take_limit: int = 100,
    client: httpx.Client | None = None,
    host_last_call: dict[str, float] | None = None,
) -> list[IngestRow]:
    """Ingest Commons divisions in which the member participated.

    Two-step pattern (per sources.v2.yaml gotcha):
      1. GET /data/divisions.json/search?queryParameters.memberId=...&...&take=N
         — returns DivisionId list but Ayes:[]/Noes:[] empty.
      2. For each DivisionId: GET /data/division/{id}.json — returns full
         per-member membership lists. Look up the member in Ayes / Noes /
         NoVoteRecorded to determine vote_position.
    """
    metrics.name = "starmer_commons_votes"
    t0 = time.monotonic()
    host_last_call = host_last_call if host_last_call is not None else {}
    own_client = client is None
    client = client or httpx.Client()

    try:
        if end_date is None:
            end_date = datetime.now(timezone.utc)
        if start_date is None:
            start_date = end_date - timedelta(days=DEFAULT_LOOKBACK_DAYS)

        search_url = (
            f"{VOTES_SEARCH}?queryParameters.memberId={member_id}"
            f"&queryParameters.startDate={start_date.strftime('%Y-%m-%d')}"
            f"&queryParameters.endDate={end_date.strftime('%Y-%m-%d')}"
            f"&queryParameters.take={take_limit}"
        )
        listing = _http_get_json(client, search_url, metrics, host_last_call)
        if listing is None:
            metrics.fatal_error = "Commons Votes search returned no data"
            return []

        # Listing is a list-or-dict depending on API revision; normalise.
        divisions = listing if isinstance(listing, list) else listing.get("Divisions") or []
        log.info("[commons-votes] %d divisions in search window", len(divisions))

        rows: list[IngestRow] = []
        for div in divisions:
            division_id = div.get("DivisionId")
            if not division_id:
                continue
            detail_url = VOTES_DETAIL.format(division_id=division_id)
            detail = _http_get_json(client, detail_url, metrics, host_last_call)
            if not detail:
                continue

            vote_position = _resolve_vote_position(detail, member_id)
            if vote_position is None:
                # Member was eligible but no record — skip rather than store ambiguous.
                continue

            occurred = _parse_iso_dt(detail.get("Date") or div.get("Date")) or end_date
            title = (detail.get("Title") or f"Division {division_id}")[:300]
            summary = (
                f"Vote position: {vote_position}. "
                f"Ayes: {detail.get('AyeCount', 0)}, Noes: {detail.get('NoCount', 0)}."
            )
            source_url = (
                f"https://commonsvotes.digiminster.com/Divisions/Details/{division_id}"
            )

            payload = {
                "DivisionId": division_id,
                "Date": detail.get("Date"),
                "Number": detail.get("Number"),
                "Title": detail.get("Title"),
                "AyeCount": detail.get("AyeCount"),
                "NoCount": detail.get("NoCount"),
                "vote_position": vote_position,
                "member_id": member_id,
            }

            rows.append(IngestRow(
                source_id="commons-votes-divisions",
                tier="T1",
                action_class="vote",
                subject_role="voter",
                ext_id=str(division_id),
                occurred_at=occurred,
                title=title,
                summary=summary,
                source_url=source_url,
                payload_json=payload,
            ))
        metrics.rows_yielded = len(rows)
        return rows
    except Exception as e:
        metrics.fatal_error = f"commons-votes strategy crashed: {type(e).__name__}: {e}"
        log.exception("commons-votes strategy crash")
        return []
    finally:
        metrics.duration_sec = time.monotonic() - t0
        if own_client:
            client.close()


def _resolve_vote_position(detail: dict[str, Any], member_id: int) -> str | None:
    """Look the member up in Ayes / Noes / NoVoteRecorded lists.

    Returns 'aye' | 'no' | 'no_vote_recorded' | None.
    """
    def has_member(lst: list[dict[str, Any]] | None) -> bool:
        if not lst:
            return False
        return any((m.get("MemberId") == member_id) for m in lst)

    if has_member(detail.get("Ayes")):
        return "aye"
    if has_member(detail.get("Noes")):
        return "no"
    if has_member(detail.get("NoVoteRecorded")):
        return "no_vote_recorded"
    return None


# ===========================================================================
# Strategy 3 — UK Lobbying Register quarterly XLSX  (lobby_filing / lobbied)
# ===========================================================================

LOBBY_PUBLICATIONS_INDEX = "https://registrarofconsultantlobbyists.org.uk/publications/"

# Names that mean Starmer is the lobbied party (broaden over time as data shows up).
# Per Phase B finding: register lists "lobbied person" by name + office.
STARMER_LOBBIED_NAMES = [
    "keir starmer",
    "sir keir starmer",
    "rt hon sir keir starmer",
    "office of the prime minister",
    "prime minister",
    "pm's office",
    "pm office",
]


def strategy_starmer_lobbying_xlsx(
    metrics: StrategyMetrics,
    *,
    quarters_back: int = 4,
    client: httpx.Client | None = None,
    host_last_call: dict[str, float] | None = None,
) -> list[IngestRow]:
    """Download recent quarterly lobby register XLSXs and yield filings naming Starmer.

    Per Phase B finding: every filing returned that mentions Starmer (or his
    office) is one where he is the LOBBIED party — per addendum DR
    subject_role=lobbied.

    NOT an API. We:
      1. GET the publications page with browser UA → parse for `*.xlsx` links.
      2. Take the most recent N quarterly XLSXs.
      3. For each: download, parse with openpyxl, filter rows by lobbied-name.
      4. Yield one IngestRow per filing.
    """
    metrics.name = "starmer_lobbying_xlsx"
    t0 = time.monotonic()
    host_last_call = host_last_call if host_last_call is not None else {}
    own_client = client is None
    client = client or httpx.Client()

    try:
        # 2026-05-12 live-ingest fix: ORCL site reorganised — the /publications/
        # index no longer hosts XLSXs directly. It now lists per-quarter article
        # URLs (e.g. /quarter-four-january-march-2026returns-data/), and each
        # article page contains the actual .xlsx download. Two-stage scrape:
        #   1. /publications/ → list of /quarter-*returns-data/ article URLs
        #   2. each article  → .xlsx download URL under /wp-content/uploads/
        from urllib.parse import urljoin
        index_bytes = _http_get_bytes(client, LOBBY_PUBLICATIONS_INDEX, metrics, host_last_call)
        if not index_bytes:
            metrics.fatal_error = "lobby register publications page unreachable"
            return []
        index_html = index_bytes.decode("utf-8", errors="replace")

        # Stage 1: extract per-quarter article URLs from the publications page.
        # Pattern: https://registrarofconsultantlobbyists.org.uk/quarter-*returns-data*/
        # Dedupe (the page repeats each href).
        article_hrefs = re.findall(
            r'href=["\'](https?://registrarofconsultantlobbyists\.org\.uk/quarter-[^"\']*?returns-data[^"\']*?/)["\']',
            index_html,
            flags=re.IGNORECASE,
        )
        seen: set[str] = set()
        article_urls: list[str] = []
        for u in article_hrefs:
            if u not in seen:
                seen.add(u)
                article_urls.append(u)
        log.info("[lobby] discovered %d quarterly article URLs on /publications/", len(article_urls))

        # Stage 2: for each article, fetch and extract .xlsx hrefs.
        xlsx_urls: list[str] = []
        for art_url in article_urls:
            art_bytes = _http_get_bytes(client, art_url, metrics, host_last_call)
            if not art_bytes:
                continue
            art_html = art_bytes.decode("utf-8", errors="replace")
            found = re.findall(
                r'href=["\']([^"\']+\.xlsx)["\']', art_html, flags=re.IGNORECASE
            )
            for href in found:
                full = urljoin(art_url, href)
                if full not in xlsx_urls and "/wp-content/uploads/" in full:
                    xlsx_urls.append(full)

        # Sort descending by URL (newer year folders sort later lexicographically).
        xlsx_urls.sort(reverse=True)
        targets = xlsx_urls[:quarters_back]
        log.info("[lobby] discovered %d XLSX URLs across articles; using %d", len(xlsx_urls), len(targets))

        rows: list[IngestRow] = []
        for xlsx_url in targets:
            xlsx_bytes = _http_get_bytes(client, xlsx_url, metrics, host_last_call)
            if not xlsx_bytes:
                continue
            try:
                filings = _parse_lobby_xlsx(xlsx_bytes, xlsx_url)
            except Exception as e:
                log.warning("lobby XLSX parse failure on %s: %s", xlsx_url, e)
                continue
            for f in filings:
                if not _filing_matches_starmer(f):
                    continue
                ext_id = _filing_ext_id(f, xlsx_url)
                occurred = f.get("_period_end_dt") or datetime.now(timezone.utc)
                title = (
                    f"Lobby filing: {f.get('lobbyist_org', '?')} -> "
                    f"{f.get('lobbied_official', 'PM office')}"
                )[:300]
                summary = (f.get("activity_description") or "")[:600]
                rows.append(IngestRow(
                    source_id="lobbying-register-quarterly-xlsx",
                    tier="T1",
                    action_class="lobby_filing",
                    subject_role="lobbied",  # per addendum DR default
                    ext_id=ext_id,
                    occurred_at=occurred,
                    title=title,
                    summary=summary,
                    source_url=xlsx_url,
                    payload_json=f,
                ))
        metrics.rows_yielded = len(rows)
        return rows
    except Exception as e:
        metrics.fatal_error = f"lobby strategy crashed: {type(e).__name__}: {e}"
        log.exception("lobby strategy crash")
        return []
    finally:
        metrics.duration_sec = time.monotonic() - t0
        if own_client:
            client.close()


def _parse_lobby_xlsx(xlsx_bytes: bytes, source_url: str) -> list[dict[str, Any]]:
    """Parse the quarterly register XLSX into a list of filing dicts.

    2026-05-12 finding: the current public ORCL quarterly XLSX format
    (verified against Q3 2025, Q4 2025, Q1 2026 downloads) contains only two
    columns: "Registered organisation name" and "Client name". It does NOT
    include a lobbied-official column, an activity description, or filing
    dates beyond the quarter named in the worksheet title. This means
    Starmer-by-name filtering cannot be applied against this dataset at all
    — the dataset simply does not name lobbied parties.

    The parser still extracts what it can (registrant + client + the period
    inferred from the sheet title), but `_filing_matches_starmer` will
    correctly yield 0 rows. The Phase B premise — register names lobbied PM
    — was based on an earlier XLSX format. See result.md in
    `Vadim's Inbox/builder-hansard-lobbying-fixes-2026-05-12/`.

    Header row position varies between quarters: some XLSXs put a title in
    row 0 plus a help-text row in row 1, with the header on row 2; others
    put the title in row 0 and the header on row 1. Scan the first 5 rows
    for the row containing the canonical "registered organisation name" /
    "client name" header pair.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        log.error("openpyxl not installed — cannot parse lobby register XLSX")
        return []

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.active

    # Try to infer the period covered from the sheet title (e.g. "Quarterly
    # Return - January to March 2026") so each filing carries an occurred_at.
    sheet_title_period_dt: datetime | None = None
    first_rows: list[tuple[Any, ...]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        first_rows.append(row)
        if i >= 30:
            break
    if first_rows:
        title_cell = first_rows[0][0] if first_rows[0] else None
        if isinstance(title_cell, str):
            m = re.search(r"(20\d{2})", title_cell)
            if m:
                sheet_title_period_dt = datetime(int(m.group(1)), 12, 31, tzinfo=timezone.utc)

    # Locate the header row.
    header_idx = -1
    norm_headers: list[str] = []
    for i, row in enumerate(first_rows):
        if not row:
            continue
        cells_low = [str(c).strip().lower() if c is not None else "" for c in row]
        if "registered organisation name" in cells_low and "client name" in cells_low:
            header_idx = i
            norm_headers = [_normalise_lobby_header(c) for c in row]
            break
        # Legacy header pattern.
        if any("lobbyist" in c for c in cells_low):
            header_idx = i
            norm_headers = [_normalise_lobby_header(c) for c in row]
            break

    if header_idx < 0:
        log.warning(
            "lobby XLSX: no recognisable header row in first 30 rows of %s", source_url
        )
        return []

    # Build full row iterator restarting after header.
    filings: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        rec: dict[str, Any] = {}
        for i, val in enumerate(row):
            if i >= len(norm_headers):
                break
            rec[norm_headers[i]] = val
        # Period-end heuristic: prefer an in-row date column, else fall back
        # to the quarter inferred from the sheet title.
        for key in ("period_covered_end", "period_end", "quarter_end", "period_covered_to"):
            if rec.get(key):
                try:
                    if isinstance(rec[key], datetime):
                        rec["_period_end_dt"] = rec[key].replace(tzinfo=timezone.utc)
                    else:
                        rec["_period_end_dt"] = _parse_iso_dt(str(rec[key]))
                except Exception:
                    pass
                break
        if "_period_end_dt" not in rec and sheet_title_period_dt is not None:
            rec["_period_end_dt"] = sheet_title_period_dt
        rec["_source_xlsx"] = source_url
        filings.append(rec)
    return filings


def _normalise_lobby_header(h: str) -> str:
    """Map register column-name variants to canonical short keys."""
    h_low = (h or "").strip().lower()
    if not h_low:
        return "unknown"
    # 2026 format uses "Registered organisation name" for the lobbyist/registrant.
    if h_low == "registered organisation name":
        return "lobbyist_org"
    if (("lobbyist" in h_low and "name" in h_low)
            or h_low in ("registrant", "registered lobbyist", "lobbyist")):
        return "lobbyist_org"
    if h_low == "client name" or "client" in h_low:
        return "client"
    if "activity" in h_low or "description" in h_low:
        return "activity_description"
    if "lobbied" in h_low and ("person" in h_low or "official" in h_low or "name" in h_low):
        return "lobbied_official"
    if "lobbied" in h_low and ("office" in h_low or "department" in h_low):
        return "lobbied_office"
    if "period" in h_low and ("end" in h_low or "to" in h_low):
        return "period_covered_end"
    if "period" in h_low and ("start" in h_low or "from" in h_low):
        return "period_covered_start"
    if "quarter" in h_low:
        return "quarter"
    if "uid" in h_low or "id" in h_low or "reference" in h_low:
        return "filing_uid"
    # Fallback: snake_case the header.
    return re.sub(r"[^a-z0-9]+", "_", h_low).strip("_")


def _filing_matches_starmer(f: dict[str, Any]) -> bool:
    haystack = " ".join(
        str(f.get(k, ""))
        for k in ("lobbied_official", "lobbied_office", "activity_description", "client")
    ).lower()
    return any(name in haystack for name in STARMER_LOBBIED_NAMES)


def _filing_ext_id(f: dict[str, Any], source_url: str) -> str:
    """Stable ID per filing: prefer the register's UID column; else hash."""
    uid = f.get("filing_uid")
    if uid:
        return str(uid)
    # Fallback: hash of identifying fields + source XLSX URL.
    identifier_parts = [
        str(f.get("lobbyist_org", "")),
        str(f.get("client", "")),
        str(f.get("lobbied_official", "")),
        str(f.get("period_covered_end", "")),
        str(f.get("activity_description", "") or "")[:200],
        source_url,
    ]
    return hashlib.sha256("|".join(identifier_parts).encode("utf-8")).hexdigest()[:32]


# ===========================================================================
# Strategy 4 — Members API  (supporting_metadata)
# ===========================================================================

MEMBERS_API_SEARCH = "https://members-api.parliament.uk/api/Members/Search"
MEMBERS_API_DETAIL = "https://members-api.parliament.uk/api/Members/{member_id}"


def strategy_starmer_members_api(
    member_id: int,
    metrics: StrategyMetrics,
    *,
    client: httpx.Client | None = None,
    host_last_call: dict[str, float] | None = None,
) -> list[IngestRow]:
    """Resolve Starmer's identity/role record. One IngestRow per call.

    Per addendum DR: action_class=supporting_metadata — decorative only, never
    load-bearing for verdict computation. The scorecard renderer reads this
    row's payload for the header (constituency, party, role).
    """
    metrics.name = "starmer_members_api"
    t0 = time.monotonic()
    host_last_call = host_last_call if host_last_call is not None else {}
    own_client = client is None
    client = client or httpx.Client()

    try:
        detail_url = MEMBERS_API_DETAIL.format(member_id=member_id)
        data = _http_get_json(client, detail_url, metrics, host_last_call)
        if not data:
            metrics.fatal_error = "Members API detail returned no data"
            return []

        # Some Members API revisions wrap the record in {"value": {...}}.
        value = data.get("value") if isinstance(data, dict) and "value" in data else data

        now = datetime.now(timezone.utc)
        name = value.get("nameDisplayAs") or value.get("nameFullTitle") or "Sir Keir Starmer"
        title = f"Identity: {name}"
        latest_house = value.get("latestHouseMembership") or {}
        constituency = latest_house.get("membershipFrom") or "Holborn and St Pancras"
        party = (value.get("latestParty") or {}).get("name") or "Labour"
        summary = f"{name} — {party} — {constituency}"

        ext_id = f"members-api-{member_id}-{now.date().isoformat()}"

        row = IngestRow(
            source_id="members-api-membership",
            tier="T1",
            action_class="supporting_metadata",
            subject_role=None,
            ext_id=ext_id,
            occurred_at=now,
            title=title[:300],
            summary=summary[:600],
            source_url=detail_url,
            payload_json=value if isinstance(value, dict) else {"raw": data},
        )
        metrics.rows_yielded = 1
        return [row]
    except Exception as e:
        metrics.fatal_error = f"members-api strategy crashed: {type(e).__name__}: {e}"
        log.exception("members-api strategy crash")
        return []
    finally:
        metrics.duration_sec = time.monotonic() - t0
        if own_client:
            client.close()


# ---------------------------------------------------------------------------
# Registry: slug → strategy callable (for the pipeline driver)
# ---------------------------------------------------------------------------

# Each entry: source_id → callable that returns (rows, metrics).
STARMER_STRATEGIES: dict[str, str] = {
    "hansard-search-contributions": "strategy_starmer_hansard",
    "commons-votes-divisions": "strategy_starmer_commons_votes",
    "lobbying-register-quarterly-xlsx": "strategy_starmer_lobbying_xlsx",
    "members-api-membership": "strategy_starmer_members_api",
}
